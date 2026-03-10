from pathlib import Path

from openpyxl import load_workbook

from app.services.adapters.base import RawCameraRecord, SourceAdapter


class ExcelAdapter(SourceAdapter):
    async def fetch(self, config: dict) -> list[RawCameraRecord]:
        file_path = Path(config["file_path"])
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        column_mapping = config.get("column_mapping", {})
        type_mapping = config.get("type_mapping", {})
        sheet_name = config.get("sheet_name", None)

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 2:
            return []

        headers = [str(h).strip() if h else "" for h in rows[0]]
        records = []
        for row in rows[1:]:
            row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
            records.append(self._parse_row(row_dict, column_mapping, type_mapping))

        return records

    def _parse_row(self, row: dict, mapping: dict, type_mapping: dict | None = None) -> RawCameraRecord:
        lat = self._parse_float(self._get_mapped(row, mapping, "lat"))
        lon = self._parse_float(self._get_mapped(row, mapping, "lon"))
        raw_type = self._get_mapped(row, mapping, "type") or "fixed_speed"
        camera_type = self._map_type(str(raw_type), type_mapping) if raw_type else "fixed_speed"
        speed_limit = self._parse_int(self._get_mapped(row, mapping, "speed_limit"))
        heading = self._parse_float(self._get_mapped(row, mapping, "heading"))
        address = self._get_mapped(row, mapping, "address") or None
        external_id = self._get_mapped(row, mapping, "external_id") or None

        return RawCameraRecord(
            lat=lat,
            lon=lon,
            type=camera_type,
            speed_limit=speed_limit,
            heading=heading,
            address=address,
            external_id=str(external_id) if external_id is not None else None,
            raw_data={k: v for k, v in row.items()},
        )

    def _map_type(self, raw_type: str, type_mapping: dict | None) -> str:
        if not type_mapping:
            return raw_type
        for key, value in type_mapping.items():
            if key in raw_type:
                return value
        return raw_type

    def _get_mapped(self, row: dict, mapping: dict, field: str):
        column_name = mapping.get(field, field)
        value = row.get(column_name)
        if value is None:
            return None
        if isinstance(value, str) and not value.strip():
            return None
        return value

    def _parse_int(self, value) -> int | None:
        if value is None:
            return None
        try:
            return int(value)
        except (ValueError, TypeError):
            return None

    def _parse_float(self, value) -> float | None:
        if value is None:
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None
